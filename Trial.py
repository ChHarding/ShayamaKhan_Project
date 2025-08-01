import tkinter as tk
from tkinter import filedialog, messagebox
import os
import fitz
import pandas as pd
import openpyxl
from openai import OpenAI
from keys import open_ai_api_key

# Initialize OpenAI client
client = OpenAI(api_key=open_ai_api_key)

# ----------- Paragraph Extraction -----------

def extract_paragraphs(text, client):
    prompt = (
        "Split the following text into paragraphs. "
        "After each paragraph, output the marker <P>.\n\n"
        + text
    )
    paragraphs = []
    buffer = ""

    with client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        stream=True,
        temperature=0
    ) as stream:
        for chunk in stream:
            if hasattr(chunk, "choices") and chunk.choices:
                delta = chunk.choices[0].delta
                if hasattr(delta, "content") and delta.content:
                    buffer += delta.content

    while "<P>" in buffer:
        para, buffer = buffer.split("<P>", 1)
        para = para.strip()
        if para:
            paragraphs.append(para)

    return paragraphs

# ----------- Tone Classification -----------

def classify_tone(paragraph, keyword, client):
    prompt = (
        f"Classify the tone of the following paragraph with respect to the keyword: {keyword}\n\n"
        "There are three possible tones:\n"
        "1. Supportive: The paragraph affirms or supports the presence of an effect, relationship, or influence of the keyword.\n"
        "2. Neutral: The paragraph mentions the keyword without taking a stance.\n"
        "3. Opposing: The paragraph indicates no effect or contradictory evidence.\n\n"
        f"Paragraph: {paragraph}\n\n"
        "Respond with: Supportive, Neutral, or Opposing, and explain why."
    )
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You are a helpful academic tone analysis assistant."},
            {"role": "user", "content": prompt}
        ],
        temperature=0
    )
    return response.choices[0].message.content.strip()

# ----------- Main Execution -----------

def run_analysis(pdf_path, keywords):
    print("Starting analysis...")  # Add this
    doc = fitz.open(pdf_path)
    all_paragraphs = []
    for page in doc:
        page_text = page.get_text("text")
        paragraphs = extract_paragraphs(page_text, client)
        all_paragraphs.extend(paragraphs)
    print(f"Extracted {len(all_paragraphs)} paragraphs.")  # Add this

    tone_score = {"Supportive": 1, "Neutral": 0, "Opposing": -1}
    excel_rows = []
    overall_tones = []
    row_ranges = []

    for keyword in keywords:
        keyword_paragraphs = [p for p in all_paragraphs if keyword.lower() in p.lower()]
        scores = []
        start_row = len(excel_rows) + 1  # +1 for header (Excel is 1-based)
        if not keyword_paragraphs:
            # Add a "not discussed" row
            excel_rows.append([
                keyword.capitalize(),
                "-",
                f"This paper does not discuss '{keyword}'.",
                "Not Discussed",
                "The keyword was not found in the document.",
                0,
                0,
                "Not Discussed"
            ])
            end_row = len(excel_rows)  # Only one row added
            row_ranges.append((start_row, end_row))
            overall_tones.append("Not Discussed")
            continue
        for i, para in enumerate(keyword_paragraphs):
            tone_full = classify_tone(para, keyword, client)
            if "." in tone_full:
                tone, explanation = tone_full.split(".", 1)
                tone = tone.strip()
                explanation = explanation.strip()
            else:
                tone = tone_full.strip()
                explanation = ""
            score = tone_score.get(tone, 0)
            scores.append(score)
            cumulative = sum(scores)
            excel_rows.append([
                keyword.capitalize(),
                f"Paragraph {i+1}",
                para,           # Extracted Text
                tone,
                explanation,    # New Explanation column
                score,
                cumulative,
                ""  # Placeholder for Overall Tone
            ])
        # Calculate overall tone for this keyword
        avg_score = sum(scores) / len(scores) if scores else 0
        if avg_score > 0:
            overall = "Supportive"
        elif avg_score < 0:
            overall = "Opposing"
        else:
            overall = "Neutral"
        # Store the range of rows for this keyword to merge later
        end_row = len(excel_rows)  # End row is the last row for this keyword
        row_ranges.append((start_row, end_row))
        overall_tones.append(overall)

    # Fill in the Overall Tone column for each keyword's rows
    for (start, end), overall in zip(row_ranges, overall_tones):
        for i in range(start-1, end):  # start-1 to end-1 inclusive
            excel_rows[i][7] = overall  # 7 is the 8th column (Overall Tone)

    df = pd.DataFrame(
        excel_rows,
        columns=[
            "Keyword",
            "Paragraph #",
            "Extracted Text",
            "Tone",
            "Explanation",
            "Score",
            "Cumulative Score",
            "Overall Tone"
        ]
    )
    output_excel = "factor_analysis_results.xlsx"
    if os.path.exists(output_excel):
        try:
            os.remove(output_excel)
            print("Old Excel file removed.")  # Add this
        except PermissionError:
            print("Excel file is open. Cannot overwrite.")  # Add this
            messagebox.showerror(
                "Error",
                f"Cannot overwrite '{output_excel}'.\nPlease close the file in Excel and try again."
            )
            return
    print("Saving new results to Excel...")  # Add this
    df.to_excel(output_excel, index=False)
    wb = openpyxl.load_workbook(output_excel)
    ws = wb.active

    # Merge cells ONLY for Overall Tone column (column 8), not for Keyword
    for (start, end) in row_ranges:
        if end - start > 1:  # Only merge if more than one row for the keyword
            ws.merge_cells(start_row=start+1, start_column=8, end_row=end+1, end_column=8)  # Overall Tone

    # --- Formatting for cleaner Excel output ---
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)  # Limit width

    # Wrap text for Extracted Text and Explanation columns
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):  # Columns C, D, E
        for cell in row:
            cell.alignment = wrap_alignment

    # Header formatting
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add borders to all cells
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(output_excel)
    print("Excel file updated successfully.")  # Add this

# --- GUI code ---

def browse_pdf():
    filename = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    pdf_path_var.set(filename)

def start_analysis():
    pdf_path = pdf_path_var.get()
    keywords = [k.strip() for k in keywords_var.get().split(",") if k.strip()]
    if not pdf_path or not keywords:
        messagebox.showerror("Error", "Please provide all inputs.")
        return
    try:
        run_analysis(pdf_path, keywords)
        messagebox.showinfo("Done", "Analysis complete! Results saved to factor_analysis_results.xlsx")
    except Exception as e:
        messagebox.showerror("Error", str(e))

root = tk.Tk()
root.title("Qualitative Data Analyzer")

pdf_path_var = tk.StringVar()
keywords_var = tk.StringVar()

tk.Label(root, text="PDF File:").grid(row=0, column=0, sticky="e")
tk.Entry(root, textvariable=pdf_path_var, width=40).grid(row=0, column=1)
tk.Button(root, text="Browse", command=browse_pdf).grid(row=0, column=2)

tk.Label(root, text="Keywords (comma separated):").grid(row=1, column=0, sticky="e")
tk.Entry(root, textvariable=keywords_var, width=40).grid(row=1, column=1, columnspan=2)

tk.Button(root, text="Run Analysis", command=start_analysis).grid(row=2, column=1, pady=10)

root.mainloop()

